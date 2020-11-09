#!/usr/bin/env python

'''Schreiben der Daten in die Excel-Datei und Formatieren'''

__author__ = 'Andreas Iakobou'
__version__ = '1.0'
__email__ = 'andreas.iakobou@partner.bmw.de'

import datetime
from datetime import timedelta
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import colors
from openpyxl.styles import Font, Color

def prepare_excel(monday, workbook, sheet, Prüffeld, KW, year):

	week = list()
	for i in range(-1, 6):
		week.append(monday+timedelta(days=i))

	if Prüffeld == 'EETZ':
		sheet['A1'] = 'KW{}'.format(str(KW))
		sheet['A2'] = '{}'.format(year)
	if Prüffeld == 'UPZ':
		sheet['D1'] = 'KW{}'.format(str(KW))
		sheet['A1'] = '{}'.format(year)
	for i, j in zip(['A4','A6','A19','A32','A45','A58','A71'], week):
		sheet[i] = j.strftime("%d.%m")


def write_data(workbook, sheet, roww, columnn, akf, colorcode, Prüffeld):
	if Prüffeld == 'UPZ':
		limit = 39
	elif Prüffeld == 'EETZ':
		limit = 60

	while True:
		if sheet.cell(row = roww, column = columnn).value is None and sheet.cell(row = roww+1, column = columnn).value is None:
			try:
				sheet.cell(row = roww, column = columnn).value = akf[4]
				sheet.cell(row = roww+1, column = columnn).value = akf[3]
				sheet.cell(row = roww, column = columnn + 4).value = akf[1]
				sheet.cell(row = roww+1, column = columnn + 4).value = akf[2]
				akf.append('planned')
				break
			except:
				print('Fehler: AKF mit V-Nr.{} vom {} konnte nicht eingeplant werden'.format(akf[4],akf[1]))
				break
		else:
			if columnn < limit:
				columnn+=7
			else:
				columnn = 4
				roww+=2

	greyfill = PatternFill("solid", start_color='DDDDDDDD')
	greenfill = PatternFill("solid", start_color= 'c5e0b4')
	bluefill = PatternFill("solid", start_color= 'dae3f3')

	if colorcode == 'grey':
		sheet.cell(row = roww, column = columnn).font = Font(bold = True)
		sheet.cell(row = roww, column = columnn).fill =  greyfill
		sheet.cell(row = roww+1, column = columnn).fill =  greyfill
		sheet.cell(row = roww, column = columnn + 4).fill =  greyfill
		sheet.cell(row = roww+1, column = columnn + 4).fill =  greyfill

	elif colorcode == 'ctk':
		sheet.cell(row = roww, column = columnn).font = Font(bold = True)
		sheet.cell(row = roww, column = columnn).fill =  greenfill
		sheet.cell(row = roww+1, column = columnn).fill =  greenfill
		sheet.cell(row = roww, column = columnn + 4).fill =  greenfill
		sheet.cell(row = roww+1, column = columnn + 4).fill =  greenfill

	elif colorcode == 'b':
		sheet.cell(row = roww, column = columnn).font = Font(bold = True)
		sheet.cell(row = roww, column = columnn).fill =  bluefill
		sheet.cell(row = roww+1, column = columnn).fill =  bluefill
		sheet.cell(row = roww, column = columnn + 4).fill =  bluefill
		sheet.cell(row = roww+1, column = columnn + 4).fill =  bluefill
	else:
		pass
